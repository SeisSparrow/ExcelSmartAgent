"""
Excel Preprocessor Module
Handles complex Excel file processing and reshaping into 2D tables
"""
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import openpyxl
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Process and reshape Excel files into standardized 2D tables"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.metadata = {}
        
    def load_file(self) -> Dict[str, pd.DataFrame]:
        """
        Load Excel file and return a dictionary of DataFrames (one per sheet)
        """
        try:
            # Try reading with openpyxl (for .xlsx)
            if self.file_path.suffix == '.xlsx':
                self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
                sheets = {}
                for sheet_name in self.workbook.sheetnames:
                    df = self._process_sheet(sheet_name)
                    if df is not None and not df.empty:
                        sheets[sheet_name] = df
                return sheets
            # For .xls or .csv
            elif self.file_path.suffix == '.xls':
                return pd.read_excel(self.file_path, sheet_name=None, engine='xlrd')
            elif self.file_path.suffix == '.csv':
                return {'Sheet1': pd.read_csv(self.file_path)}
            else:
                raise ValueError(f"Unsupported file format: {self.file_path.suffix}")
        except Exception as e:
            logger.error(f"Error loading file {self.file_path}: {e}")
            raise
    
    def _process_sheet(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """
        Process a single sheet and handle complex structures
        """
        try:
            sheet = self.workbook[sheet_name]
            
            # Detect merged cells
            merged_ranges = self._get_merged_cells(sheet)
            
            # Find the actual data range (skip empty rows/columns)
            data_range = self._find_data_range(sheet)
            
            if not data_range:
                return None
            
            # Detect header row(s)
            header_row = self._detect_header_row(sheet, data_range)
            
            # Extract data with proper handling of merged cells
            df = self._extract_data(sheet, data_range, header_row, merged_ranges)
            
            # Clean and standardize the DataFrame
            df = self._clean_dataframe(df)
            
            # Store metadata
            self.metadata[sheet_name] = {
                'original_range': data_range,
                'header_row': header_row,
                'columns': list(df.columns),
                'row_count': len(df),
                'column_count': len(df.columns)
            }
            
            return df
            
        except Exception as e:
            logger.error(f"Error processing sheet {sheet_name}: {e}")
            return None
    
    def _get_merged_cells(self, sheet) -> List[str]:
        """Get list of merged cell ranges"""
        return [str(merged_range) for merged_range in sheet.merged_cells.ranges]
    
    def _find_data_range(self, sheet) -> Optional[Tuple[int, int, int, int]]:
        """
        Find the actual data range in the sheet (min_row, max_row, min_col, max_col)
        Skips completely empty rows and columns
        """
        min_row, max_row = None, None
        min_col, max_col = None, None
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    if min_row is None or cell.row < min_row:
                        min_row = cell.row
                    if max_row is None or cell.row > max_row:
                        max_row = cell.row
                    if min_col is None or cell.column < min_col:
                        min_col = cell.column
                    if max_col is None or cell.column > max_col:
                        max_col = cell.column
        
        if min_row is None:
            return None
            
        return (min_row, max_row, min_col, max_col)
    
    def _detect_header_row(self, sheet, data_range: Tuple[int, int, int, int]) -> int:
        """
        Detect which row contains the column headers
        Uses heuristics: headers are usually text, data below are mixed or numeric
        """
        min_row, max_row, min_col, max_col = data_range
        
        # Check first few rows
        for row_idx in range(min_row, min(min_row + 5, max_row + 1)):
            row_values = []
            for col_idx in range(min_col, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_values.append(cell.value)
            
            # If row has mostly text and next row has numbers, this is likely the header
            if self._is_likely_header(row_values):
                return row_idx
        
        # Default to first row
        return min_row
    
    def _is_likely_header(self, values: List[Any]) -> bool:
        """Check if a row is likely to be a header row"""
        non_empty = [v for v in values if v is not None]
        if not non_empty:
            return False
        
        text_count = sum(1 for v in non_empty if isinstance(v, str))
        return text_count / len(non_empty) > 0.7  # 70% or more are strings
    
    def _extract_data(self, sheet, data_range: Tuple[int, int, int, int], 
                     header_row: int, merged_ranges: List[str]) -> pd.DataFrame:
        """
        Extract data from sheet with proper handling of merged cells
        """
        min_row, max_row, min_col, max_col = data_range
        
        # Extract headers
        headers = []
        for col_idx in range(min_col, max_col + 1):
            cell = sheet.cell(row=header_row, column=col_idx)
            value = cell.value if cell.value is not None else f"Column_{col_idx}"
            headers.append(str(value).strip())
        
        # Handle duplicate headers
        headers = self._make_unique_headers(headers)
        
        # Extract data rows
        data = []
        for row_idx in range(header_row + 1, max_row + 1):
            row_data = []
            for col_idx in range(min_col, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value)
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        return df
    
    def _make_unique_headers(self, headers: List[str]) -> List[str]:
        """Make column headers unique by appending numbers to duplicates"""
        seen = {}
        unique_headers = []
        
        for header in headers:
            if header in seen:
                seen[header] += 1
                unique_headers.append(f"{header}_{seen[header]}")
            else:
                seen[header] = 0
                unique_headers.append(header)
        
        return unique_headers
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize the DataFrame"""
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Reset index
        df = df.reset_index(drop=True)
        
        # Strip whitespace from string columns
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        return df
    
    def get_column_info(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        """
        Get detailed information about each column
        """
        column_info = {}
        
        for col in df.columns:
            column_info[col] = {
                'dtype': str(df[col].dtype),
                'non_null_count': int(df[col].notna().sum()),
                'null_count': int(df[col].isna().sum()),
                'unique_count': int(df[col].nunique()),
                'sample_values': df[col].dropna().head(5).tolist()
            }
            
            # Add statistics for numeric columns
            if pd.api.types.is_numeric_dtype(df[col]):
                column_info[col].update({
                    'mean': float(df[col].mean()) if not df[col].isna().all() else None,
                    'min': float(df[col].min()) if not df[col].isna().all() else None,
                    'max': float(df[col].max()) if not df[col].isna().all() else None,
                })
        
        return column_info
    
    def process_and_save(self, output_dir: Path) -> Dict[str, Any]:
        """
        Process the Excel file and save processed data
        Returns metadata about processed sheets
        """
        sheets = self.load_file()
        result = {
            'file_name': self.file_path.name,
            'sheets': {}
        }
        
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        for sheet_name, df in sheets.items():
            # Save processed data
            output_file = output_dir / f"{self.file_path.stem}_{sheet_name}.csv"
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
            
            # Get column info
            column_info = self.get_column_info(df)
            
            result['sheets'][sheet_name] = {
                'output_file': str(output_file),
                'row_count': len(df),
                'column_count': len(df.columns),
                'columns': list(df.columns),
                'column_info': column_info,
                'preview': df.head(3).to_dict('records')
            }
        
        return result


def process_excel_file(file_path: str, output_dir: str) -> Dict[str, Any]:
    """
    Convenience function to process an Excel file
    """
    processor = ExcelProcessor(file_path)
    return processor.process_and_save(Path(output_dir))

